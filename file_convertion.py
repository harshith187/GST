from collections import OrderedDict
from datetime import datetime
from operator import itemgetter

import openpyxl
import simplejson as json
import xlwt

def file_convertion(filepath):
    in_w_book = openpyxl.load_workbook(filename=filepath)
    in_w_sheet = in_w_book["Query Report"]
    gstin = in_w_sheet.cell(row = 2,  column = 7).value
    out_w_book = openpyxl.load_workbook(filename="/Users/harshith/projects/gst/GSTR1_Excel_Workbook_Template_V1.5.xlsx")
    row_count = in_w_sheet.max_row
    col_count = in_w_sheet.max_column
    converted_sheet = convert_to_exel(in_w_sheet,  out_w_book,  row_count,  col_count)
    converted_sheet_data_list = []
    for row in range(5, row_count+3):
        converted_sheet_data = []
        for column in range(1, 14):
            converted_sheet_data.append(converted_sheet.cell(row, column).value)
        converted_sheet_data_list.append(converted_sheet_data)
    print converted_sheet_data_list
    sort_converted_sheet(converted_sheet,  row_count)
    convert_to_json(converted_sheet_data_list,  gstin,  row_count+4)

def convert_to_exel(input_worksheet,  output_workbook,  row_count,  col_count):
    output_worksheet = output_workbook["b2b"]
    headers = ["Customer GSTIN",  "Customer Name",  "Invoice", "Posting Date", "Grand Total", "Place of Supply", "Reverse Charge", "skip", "Invoice Type", "E-Commerce GSTIN", "skip_rate", "Net Total"]
    index = 0
    for header in (headers):
        index += 1
        col_index = get_colunm_index(col_count, header, input_worksheet)
        if(header == "skip"):
            continue  
        if(header <> "skip_rate"):
            selected_data = copy_data(col_index , 2 , col_index , row_count ,  input_worksheet,   header)
        paste_data(index, 5, index, row_count+2, header, output_worksheet, input_worksheet,  selected_data)
    #output_workbook.save("converted_gstin1.xlsx")
    return output_worksheet

def copy_data(start_col,  start_row,  end_col,  end_row,  sheet, header):
    range_selected = []
    for i in range(start_row,  end_row + 1,  1):
        row_selected = []
        for j in range(start_col,  end_col+1,  1):
            if((sheet.cell(row = i,  column = j).value == "29" or sheet.cell(row = i,  column = j).value == "0" or  sheet.cell(row = i,  column = j).value == None) and header == "Place of Supply" ):
                row_selected.append("29-Karnataka")
            else:
                row_selected.append(sheet.cell(row = i,  column = j).value)
        range_selected.append(row_selected)

    return range_selected
    
    
def paste_data(start_col,  start_row,  end_col,  end_row, header,  output_sheet, input_sheet,  copied_data):
    count_row = 0
    for i in range(start_row, end_row+1, 1):
        count_col = 0
        for j in range(start_col, end_col+1, 1):
            if header == "Posting Date":
                invoice_date = (copied_data[count_row][count_col]).date()
                invoice_object =  datetime.strftime(invoice_date,  '%d-%b-%Y')  
                output_sheet.cell(row = i,  column = j).value = invoice_object
            elif(header == "skip_rate"):
                rate = calc_rate(input_sheet,  i-3) #i is the row of which the data is to be calculate 
                output_sheet.cell(row = i,  column = j).value = rate
            else:
                output_sheet.cell(row = i,  column = j).value = copied_data[count_row][count_col]
            count_col += 1
        count_row += 1

def calc_rate(sheet, row_index):
    Total_value = sheet.cell(row = row_index,  column = 26).value
    Tax_value = sheet.cell(row = row_index,  column = 27).value
    return int(round(Tax_value*200/Total_value))

def get_colunm_index(col_count, header, sheet):
    col_index = 0
    for col in sheet.iter_cols(min_row=1,  max_col=col_count,  max_row=1):
        for cell in col:
            col_index += 1
            if header == cell.value:
                return col_index  

def convert_to_json(sheet_data, gstin, row_count):
    sheet_details = OrderedDict()
    Date = datetime.strptime(sheet_data[0][3], '%d-%b-%Y')
    sheet_details['fp'] = str(Date.strftime("%m%Y"))
    sheet_details['gstin'] = str(gstin)
    sheet_details['hash'] = 'hash'
    sheet_details['version'] = 'GST2.2.6'
    gstins = []
    for row in sheet_data:
        gstins.append(row[0]) 
    data_list = []
    gstin_index = 0
    i = 0
    while(i < len(sheet_data)):
        gstin_details = OrderedDict()
        print sheet_data[i]
        gstin_details_list = sheet_data[i]
        gstin_details['ctin'] = gstin_details_list[0]
        invoice_list=[]
        while True:
            invoice = {} 
            invoice['inum'] = gstin_details_list[2]
            invoice_date = datetime.strptime(gstin_details_list[3], "%d-%b-%Y")
            invoice['idt'] = datetime.strftime(invoice_date,  '%d-%m-%Y')
            invoice['val'] = gstin_details_list[4]
            invoice['pos'] = gstin_details_list[5][:2]
            invoice['rchrg'] = gstin_details_list[6]
            invoice['inv_typ'] = gstin_details_list[8][:1]
            num_list = []
            num = OrderedDict()
            num['num'] = 1201
            items_details = OrderedDict()
            items_details['txval'] = gstin_details_list[11]
            items_details['rt'] = int(gstin_details_list[10])
            items_details['camt'] = ''
            items_details['samt'] = ''
            items_details['csamt'] = 0
            num['itm_det'] = items_details
            num_list.append(num)
            invoice['itms'] = num_list
            invoice_list.append(invoice)
            gstin_index += 1
            if ( gstin_index >= len(gstins)):
                break
            if (gstins[gstin_index-1] == gstins[gstin_index]):
                i += 1
                gstin_details_list = []
                gstin_details = OrderedDict()
                gstin_details_list.append(sheet_data[i])
            else:
                break
        i += 1
        gstin_details['inv'] = invoice_list
        sheet_details['b2b'] = data_list
        data_list.append(gstin_details)
                

    j = json.dumps(sheet_details)
    with open('final.json',  'w') as f:
        f.write(j) 

def sort_converted_sheet(converted_sheet, row_count):
    details = []
    for i in range(5,  row_count):
        row_details = []
        for cell in converted_sheet[i]:
            row_details.append(cell.value), 
        details.append(row_details)
    details = sorted(details,  key=itemgetter(1))
    for index_r,  row in enumerate(details):
        for index_c,  value in enumerate(row):
            converted_sheet.cell(row = index_r+5, column =  index_c+1).value =  value
